import boto3
import json
from django.conf import settings
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from django.http import JsonResponse
from .models import ExtractionModel, WaybillImage, ExtractedData
from .serializers import (
    ExtractionModelSerializer,
    WaybillImageSerializer,
    ExtractedDataSerializer,
)
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime
import os
import base64
from mistralai import Mistral
from django.shortcuts import render


# Simple view to test API
@api_view(["GET"])
def test_api(request):
    models = ExtractionModel.objects.all()
    serializer = ExtractionModelSerializer(models, many=True)
    return Response(serializer.data)


class ExtractionModelViewSet(viewsets.ModelViewSet):
    queryset = ExtractionModel.objects.all()
    serializer_class = ExtractionModelSerializer


class WaybillImageViewSet(viewsets.ModelViewSet):
    queryset = WaybillImage.objects.all()
    serializer_class = WaybillImageSerializer

    def get_text_for_cell(self, cell, blocks):
        """Get text for a cell by finding overlapping word blocks"""
        if "Text" in cell:
            return cell["Text"]

        cell_box = cell["Geometry"]["BoundingBox"]
        cell_text = []

        # Calculate cell boundaries with a small margin
        margin = 0.005  # Add a small margin to account for slight misalignments
        cell_left = cell_box["Left"] - margin
        cell_right = cell_box["Left"] + cell_box["Width"] + margin
        cell_top = cell_box["Top"] - margin
        cell_bottom = cell_box["Top"] + cell_box["Height"] + margin

        for block in blocks:
            if block["BlockType"] == "WORD":
                word_box = block["Geometry"]["BoundingBox"]
                word_left = word_box["Left"]
                word_right = word_box["Left"] + word_box["Width"]
                word_top = word_box["Top"]
                word_bottom = word_box["Top"] + word_box["Height"]

                # Check if word significantly overlaps with cell
                horizontal_overlap = word_right > cell_left and word_left < cell_right
                vertical_overlap = word_bottom > cell_top and word_top < cell_bottom

                if horizontal_overlap and vertical_overlap:
                    # Calculate overlap percentage
                    overlap_width = min(word_right, cell_right) - max(
                        word_left, cell_left
                    )
                    overlap_height = min(word_bottom, cell_bottom) - max(
                        word_top, cell_top
                    )
                    word_area = word_box["Width"] * word_box["Height"]
                    overlap_area = max(0, overlap_width) * max(0, overlap_height)

                    # If word overlaps significantly with the cell (>30% of word area)
                    if overlap_area > 0.3 * word_area:
                        cell_text.append(block["Text"])

        return " ".join(cell_text)

    def get_table_cells(self, table_block, blocks):
        """Get all cells belonging to a specific table"""
        if "Relationships" not in table_block:
            return []

        cell_ids = []
        for relationship in table_block["Relationships"]:
            if relationship["Type"] == "CHILD":
                cell_ids.extend(relationship["Ids"])

        cells = []
        for block in blocks:
            if block["Id"] in cell_ids and block["BlockType"] == "CELL":
                # Get the text for this cell by looking at overlapping word blocks
                block["Text"] = self.get_text_for_cell(block, blocks)
                cells.append(block)

        return cells

    def extract_table_data(self, table_block, blocks):
        """Extract table data and confidence scores from Textract blocks for a specific table"""
        cells = {}
        confidence_scores = {}

        # Get cells for this specific table
        table_cells = self.get_table_cells(table_block, blocks)

        # Process cells
        for cell in table_cells:
            row_idx = cell["RowIndex"]
            col_idx = cell["ColumnIndex"]

            # Store cell text with single quote prefix for non-empty text
            text = cell.get("Text", "").strip()
            if text:
                # Handle special characters and formatting
                text = text.replace('"', '""')  # Escape double quotes
                text = f"'{text}"

            # Store cell data
            cells[(row_idx, col_idx)] = text

            # Store confidence score
            confidence_scores[(row_idx, col_idx)] = cell.get("Confidence", 0)

        if not cells:
            return None

        # Find dimensions of the table
        max_row = max(pos[0] for pos in cells.keys())
        max_col = max(pos[1] for pos in cells.keys())

        # Create table data structure
        table_data = {
            "rows": [["" for _ in range(max_col)] for _ in range(max_row)],
            "confidence_scores": [[0 for _ in range(max_col)] for _ in range(max_row)],
        }

        # Fill in the data
        for (row_idx, col_idx), text in cells.items():
            table_data["rows"][row_idx - 1][col_idx - 1] = text
            table_data["confidence_scores"][row_idx - 1][col_idx - 1] = (
                confidence_scores.get((row_idx, col_idx), 0)
            )

        return table_data

    def extract_with_textract(self, image_path):
        if not settings.AWS_ACCESS_KEY_ID or not settings.AWS_SECRET_ACCESS_KEY:
            print("AWS credentials not found, returning demo data")
            return {
                "sender": {
                    "name": "AWS Textract Demo Sender",
                    "address": "123 AWS Street, AWS City, 12345",
                    "phone": "123-456-7890",
                },
                "recipient": {
                    "name": "AWS Textract Demo Recipient",
                    "address": "456 AWS Avenue, AWS City, 67890",
                    "phone": "987-654-3210",
                },
                "shipment": {
                    "tracking_number": "AWS123456789",
                    "date": "2025-03-17",
                    "weight": "2.5 kg",
                    "service_type": "Express",
                },
                "note": "This is demo data since AWS API keys are not configured.",
            }

        print(f"\nProcessing file with AWS Textract: {os.path.basename(image_path)}")
        print(f"Full path: {image_path}")

        textract = boto3.client(
            "textract",
            aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
            region_name=settings.AWS_REGION,
        )

        with open(image_path, "rb") as document:
            print("Reading document and calling Textract API...")
            response = textract.analyze_document(
                Document={"Bytes": document.read()}, FeatureTypes=["TABLES", "FORMS"]
            )
            print("Received response from Textract API")

        # Initialize the structured response
        structured_data = {
            "tables": [],
            "forms": {},
            "raw_text": "",
            "confidence_scores": {},
        }

        # Create a map of block IDs to blocks
        blocks_map = {block["Id"]: block for block in response["Blocks"]}
        print(f"Total blocks found: {len(response['Blocks'])}")

        # Process tables
        table_blocks = [
            block for block in response["Blocks"] if block["BlockType"] == "TABLE"
        ]
        print(f"Found {len(table_blocks)} tables")

        for i, table_block in enumerate(table_blocks, 1):
            print(f"\nProcessing Table {i}:")
            table_data = self.extract_table_data(table_block, response["Blocks"])
            if table_data:
                print(f"Table {i} data extracted successfully")
                print(f"Rows: {len(table_data['rows'])}")
                print(
                    f"Columns: {len(table_data['rows'][0]) if table_data['rows'] else 0}"
                )
                structured_data["tables"].append(table_data)
            else:
                print(f"No data found in Table {i}")

        # Process form fields (key-value pairs)
        form_fields = 0
        for block in response["Blocks"]:
            if block["BlockType"] == "KEY_VALUE_SET" and "KEY" in block["EntityTypes"]:
                key = self.get_text_for_cell(block, response["Blocks"])
                value = ""
                confidence = block.get("Confidence", 0)

                # Find the corresponding value
                for relationship in block.get("Relationships", []):
                    if relationship["Type"] == "VALUE":
                        for value_id in relationship["Ids"]:
                            value_block = blocks_map[value_id]
                            value = self.get_text_for_cell(
                                value_block, response["Blocks"]
                            )
                            confidence = value_block.get("Confidence", confidence)

                if key and value:
                    form_fields += 1
                    structured_data["forms"][key] = {
                        "value": value,
                        "confidence": confidence,
                    }

        print(f"\nExtracted {form_fields} form fields")

        # Collect raw text
        raw_text_lines = []
        for block in response["Blocks"]:
            if block["BlockType"] == "LINE":
                raw_text_lines.append(self.get_text_for_cell(block, response["Blocks"]))

        structured_data["raw_text"] = "\n".join(raw_text_lines)
        print(f"Extracted {len(raw_text_lines)} lines of raw text")

        print("\nExtraction completed successfully")
        return structured_data

    def extract_with_mistral(self, image_path):
        if not settings.MISTRAL_API_KEY:
            # Return dummy data for demo purposes
            return {
                "sender": {
                    "name": "Mistral Demo Sender",
                    "address": "123 Mistral Street, Mistral City, 12345",
                    "phone": "123-456-7890",
                },
                "recipient": {
                    "name": "Mistral Demo Recipient",
                    "address": "456 Mistral Avenue, Mistral City, 67890",
                    "phone": "987-654-3210",
                },
                "shipment": {
                    "tracking_number": "MISTRAL123456789",
                    "date": "2025-03-17",
                    "weight": "2.5 kg",
                    "service_type": "Express",
                },
                "note": "This is demo data since Mistral API key is not configured.",
            }

        try:
            # Initialize Mistral client
            client = Mistral(api_key=settings.MISTRAL_API_KEY)

            # Read the image file and encode it to base64
            with open(image_path, "rb") as image_file:
                image_data = base64.b64encode(image_file.read()).decode("utf-8")

            # Process the image using Mistral OCR with the correct format
            ocr_response = client.ocr.process(
                model="mistral-ocr-latest",
                document={
                    "type": "image_url",
                    "image_url": f"data:image/jpeg;base64,{image_data}",
                },
            )

            # Convert the OCR response to a dictionary
            response_dict = json.loads(ocr_response.model_dump_json())

            # Structure the data to include all OCR information
            structured_data = {
                "ocr_info": {
                    "model": response_dict.get("model", ""),
                    "usage_info": response_dict.get("usage_info", {}),
                },
                "pages": [],
            }

            # Process each page
            for page in response_dict.get("pages", []):
                page_data = {
                    "index": page.get("index", 0),
                    "dimensions": page.get("dimensions", {}),
                    "images": page.get("images", []),
                    "markdown": page.get("markdown", ""),
                }
                structured_data["pages"].append(page_data)

            # Add extracted text analysis
            extracted_text = ""
            for page in ocr_response.pages:
                extracted_text += page.markdown + "\n"

            structured_data["extracted_text"] = {
                "raw_text": extracted_text,
                "analysis": {
                    "sender": {},
                    "recipient": {},
                    "shipment": {},
                },
            }

            # Parse the extracted text to populate structured data
            lines = extracted_text.split("\n")
            for line in lines:
                line = line.strip()
                if "sender" in line.lower() or "from" in line.lower():
                    structured_data["extracted_text"]["analysis"]["sender"]["info"] = (
                        line
                    )
                elif "recipient" in line.lower() or "to" in line.lower():
                    structured_data["extracted_text"]["analysis"]["recipient"][
                        "info"
                    ] = line
                elif "tracking" in line.lower() or "waybill" in line.lower():
                    structured_data["extracted_text"]["analysis"]["shipment"][
                        "tracking_number"
                    ] = line
                elif "date" in line.lower():
                    structured_data["extracted_text"]["analysis"]["shipment"][
                        "date"
                    ] = line
                elif "weight" in line.lower():
                    structured_data["extracted_text"]["analysis"]["shipment"][
                        "weight"
                    ] = line

            return structured_data

        except Exception as e:
            print(f"Error in Mistral OCR extraction: {str(e)}")
            # Return dummy data if there's an error with the API
            return {
                "sender": {
                    "name": "Mistral Error Fallback",
                    "address": "Error occurred during extraction",
                    "phone": "N/A",
                },
                "recipient": {
                    "name": "Mistral Error Fallback",
                    "address": "Error occurred during extraction",
                    "phone": "N/A",
                },
                "shipment": {
                    "tracking_number": "ERROR123",
                    "date": datetime.now().strftime("%Y-%m-%d"),
                    "weight": "N/A",
                    "service_type": "N/A",
                },
                "error": str(e),
                "note": "This is fallback data due to an error with the Mistral API.",
            }

    @action(detail=False, methods=["post"])
    def bulk_upload(self, request):
        images = request.FILES.getlist("images")
        extraction_model_id = request.data.get("extraction_model")

        if not images:
            return Response(
                {"error": "No images provided"}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            extraction_model = ExtractionModel.objects.get(id=extraction_model_id)
            print(f"\nUsing extraction model: {extraction_model.name}")
        except ExtractionModel.DoesNotExist:
            return Response(
                {"error": "Invalid extraction model"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check for required API keys based on the selected model
        if (
            extraction_model.name.lower() == "aws textract"
            and not settings.AWS_ACCESS_KEY_ID
        ):
            return Response(
                {
                    "error": "AWS credentials are not configured. Please set AWS_ACCESS_KEY_ID and AWS_SECRET_ACCESS_KEY environment variables."
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        if extraction_model.name.lower() == "mistral" and not settings.MISTRAL_API_KEY:
            return Response(
                {
                    "error": "Mistral API key is not configured. Please set MISTRAL_API_KEY environment variable."
                },
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        # Ensure media directory exists
        media_root = os.path.join(settings.BASE_DIR, "media")
        os.makedirs(media_root, exist_ok=True)

        uploaded_images = []
        uploaded_ids = []  # Track the IDs of uploaded waybills

        print(f"\nProcessing {len(images)} images:")
        for idx, image in enumerate(images, 1):
            try:
                print(f"\n[{idx}/{len(images)}] Processing: {image.name}")

                # Create the waybill image record
                waybill_image = WaybillImage.objects.create(
                    image=image, extraction_model_id=extraction_model_id
                )
                print(f"Created waybill record. ID: {waybill_image.id}")
                print(f"Image saved to: {waybill_image.image.path}")

                # Store the ID for the download URL
                uploaded_ids.append(waybill_image.id)

                # Process the image based on the selected model
                try:
                    print(f"Starting extraction with {extraction_model.name}...")
                    if extraction_model.name.lower() == "aws textract":
                        extracted_data = self.extract_with_textract(
                            waybill_image.image.path
                        )
                        print("AWS Textract extraction completed")
                        if "tables" in extracted_data:
                            print(f"Found {len(extracted_data['tables'])} tables")
                            for i, table in enumerate(extracted_data["tables"], 1):
                                print(f"Table {i}: {len(table['rows'])} rows")
                    elif extraction_model.name.lower() == "mistral":
                        print("Calling Mistral API for extraction...")
                        extracted_data = self.extract_with_mistral(
                            waybill_image.image.path
                        )
                        print("Mistral API extraction completed")

                    # Create ExtractedData record
                    ExtractedData.objects.create(
                        waybill_image=waybill_image,
                        extracted_data=extracted_data,
                    )
                    print("Extraction data saved to database")

                    # Mark the waybill as processed
                    waybill_image.processed = True
                    waybill_image.save()
                    print(f"Waybill {waybill_image.id} marked as processed")

                    uploaded_images.append(waybill_image)

                except Exception as e:
                    print(f"Error processing image {image.name}: {str(e)}")
                    # Delete the waybill image if processing failed
                    waybill_image.delete()
                    return Response(
                        {"error": f"Error processing image {image.name}: {str(e)}"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    )

            except Exception as e:
                print(f"Error uploading image {image.name}: {str(e)}")
                return Response(
                    {"error": f"Error uploading image {image.name}: {str(e)}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

        if uploaded_images:
            # Construct the download URL with the correct endpoint
            download_url = (
                f"waybills/download_excel/?ids={','.join(map(str, uploaded_ids))}"
            )
            print(f"\nSuccessfully processed {len(uploaded_images)} images")
            print(f"Download URL: {download_url}")

            return Response(
                {
                    "message": f"Successfully uploaded {len(uploaded_images)} images",
                    "ids": uploaded_ids,
                    "download_url": download_url,
                },
                status=status.HTTP_201_CREATED,
            )
        else:
            return Response(
                {"error": "Failed to upload any images"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"])
    def download_excel(self, request):
        waybill_ids = request.query_params.get("ids", "")
        print(f"Downloading Excel for waybill IDs: {waybill_ids}")

        wb = Workbook()
        default_sheet = wb.active
        default_sheet.title = "Summary"

        # Add a header to the default sheet
        default_sheet["A1"] = "Waybill Extraction Summary"
        default_sheet["A2"] = "Generated on"
        default_sheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Get waybills
        if waybill_ids:
            ids_list = [
                int(id.strip()) for id in waybill_ids.split(",") if id.strip().isdigit()
            ]
            waybills = WaybillImage.objects.filter(id__in=ids_list)
        else:
            waybills = WaybillImage.objects.all()

        if not waybills.exists():
            default_sheet["A4"] = "No waybills found"
        else:
            # Add summary headers
            default_sheet["A4"] = "Waybills"
            default_sheet["A5"] = "ID"
            default_sheet["B5"] = "Upload Date"
            default_sheet["C5"] = "Extraction Model"
            default_sheet["D5"] = "Processed"

            # Add summary data
            row = 6
            for waybill in waybills:
                default_sheet[f"A{row}"] = waybill.id
                default_sheet[f"B{row}"] = waybill.uploaded_at.strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                default_sheet[f"C{row}"] = (
                    waybill.extraction_model.name if waybill.extraction_model else "N/A"
                )
                default_sheet[f"D{row}"] = "Yes" if waybill.processed else "No"
                row += 1

            # Process each waybill
            for waybill in waybills:
                try:
                    extracted_data = waybill.extracteddata.extracted_data

                    # Handle AWS Textract data
                    if isinstance(extracted_data, dict) and "tables" in extracted_data:
                        # Create sheets for each table
                        for table_idx, table in enumerate(extracted_data["tables"], 1):
                            sheet_name = f"Table_{table_idx}"
                            table_sheet = wb.create_sheet(title=sheet_name)

                            # Write table data
                            for row_idx, row_data in enumerate(table["rows"]):
                                for col_idx, cell_value in enumerate(row_data):
                                    # Remove single quote prefix if present
                                    if isinstance(
                                        cell_value, str
                                    ) and cell_value.startswith("'"):
                                        cell_value = cell_value[1:]
                                    table_sheet.cell(
                                        row=row_idx + 1,
                                        column=col_idx + 1,
                                        value=cell_value,
                                    )

                            # Add empty row
                            empty_row = len(table["rows"]) + 1

                            # Add confidence scores header
                            confidence_header_row = empty_row + 1
                            table_sheet.cell(
                                row=confidence_header_row,
                                column=1,
                                value="Confidence Scores % (Table Cell)",
                            )

                            # Write confidence scores
                            for row_idx, confidence_row_data in enumerate(
                                table["confidence_scores"]
                            ):
                                for col_idx, confidence in enumerate(
                                    confidence_row_data
                                ):
                                    table_sheet.cell(
                                        row=confidence_header_row + row_idx + 1,
                                        column=col_idx + 1,
                                        value=f"{confidence:.8f}",
                                    )

                        # Create info sheet for forms and raw text
                        info_sheet = wb.create_sheet(title="Forms_and_Text")

                        # Add form fields
                        info_sheet["A1"] = "Form Fields"
                        info_sheet["A2"] = "Field"
                        info_sheet["B2"] = "Value"
                        info_sheet["C2"] = "Confidence"

                        row = 3
                        for key, data in extracted_data.get("forms", {}).items():
                            info_sheet[f"A{row}"] = key
                            info_sheet[f"B{row}"] = data["value"]
                            info_sheet[f"C{row}"] = f"{data['confidence']:.8f}"
                            row += 1

                        # Add raw text
                        row += 2
                        info_sheet[f"A{row}"] = "Raw Text"
                        info_sheet[f"A{row + 1}"] = extracted_data.get("raw_text", "")

                    else:
                        # Handle non-Textract data
                        other_sheet = wb.create_sheet(title=f"Waybill_{waybill.id}")
                        other_sheet["A1"] = "Field"
                        other_sheet["B1"] = "Value"

                        if isinstance(extracted_data, dict):
                            row = 2
                            for key, value in extracted_data.items():
                                other_sheet[f"A{row}"] = key
                                other_sheet[f"B{row}"] = str(value)
                                row += 1
                        else:
                            other_sheet["A2"] = "Raw Data"
                            other_sheet["B2"] = str(extracted_data)

                except ExtractedData.DoesNotExist:
                    error_sheet = wb.create_sheet(title=f"Waybill_{waybill.id}")
                    error_sheet["A1"] = "No extracted data available"

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename=waybills_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

        wb.save(response)
        return response


def index(request):
    """A simple index view to help with debugging."""
    return render(
        request,
        "waybill/index.html",
        {
            "title": "Waybill Extractor API",
            "api_endpoints": [
                "/api/extraction-models/",
                "/api/waybills/",
                "/api/test-api/",
                "/admin/",
            ],
        },
    )
